from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
import os
import datetime
import pytz

base_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
print(base_path)
os.chdir(base_path)
file_path = os.path.join('Auto', 'Full-Auto', 'CI', 'paradox_develop.txt')
job_categories = ['先锋', '近卫', '重装', '狙击', '术师', '医疗', '辅助', '特种']
now = datetime.datetime.now()
date = now.astimezone(pytz.timezone('Asia/Shanghai')).strftime('%Y.%m.%d %H:%M:%S')


class DoTxt:
    def __init__(self, file_name):
        self.file_name = file_name

    def read_paradox_lines(self):
        data = []
        with open(self.file_name, "r", encoding="utf-8") as f:
            lines = f.readlines()
        job_num = 0
        sub_data = [[job_categories[job_num], '有无', '代码', None]]
        for line in lines:
            if line == '\n':
                data.append(sub_data)
                if job_num == 7:
                    break
                job_num += 1
                sub_data = [[job_categories[job_num], '有无', '代码', None]]
                continue
            parts = line.strip().split('\t')
            job = parts[0]
            num = parts[1]
            try:
                codes = parts[2]
            except IndexError:
                codes = ''
            sub_data.append([job, num, codes, None])
        return data

    def read_module_user_lines(self):
        data = [[f'更新日期：{date}'], ['干员', '关卡', '数量', '代码']]
        with open(self.file_name, "r", encoding="utf-8") as f:
            lines = f.readlines()
        for line in lines:
            parts = line.strip().split('\t')
            oper = parts[0]
            level = parts[1]
            num = parts[2]
            codes = parts[3]
            data.append([oper, level, num, codes])
        return data

    def read_module_develop_lines(self):
        data = [[f'更新日期：{date}'], ['干员', '关卡', '数量', '代码', '内容']]
        data2 = [['False'], ['False']]
        with open(self.file_name, "r", encoding="utf-8") as f:
            lines = f.readlines()
        i = 1
        for line in lines:
            parts = line.strip().split('\t')
            oper = parts[0]
            level = parts[1]
            num = parts[2]
            codes = parts[3]
            all_below_threshold = parts[4]
            content = module_task[i + 1][2]
            data.append([oper, level, num, codes, content])
            data2.append([all_below_threshold])
            i += 1
        return data, data2

    def read_module_task_lines(self):
        data = [[f'更新日期：{date}'], ['干员', '关卡', '内容']]
        with open(self.file_name, "r", encoding="utf-8") as f:
            lines = f.readlines()
        for line in lines:
            parts = line.strip().split('\t')
            oper = parts[0]
            level = parts[1]
            content = parts[2]
            data.append([oper, level, content])
        return data


class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(1, sheet.max_row + 1):
            sub_data = []
            for j in range(1, sheet.max_column + 1):
                sub_data.append(sheet.cell(row=i, column=j).value)
            test_data.append(sub_data)
        return test_data

    def write_data_all_center(self, data, start_column=1, wrap_text=False):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]

        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, value in enumerate(row_data, start=start_column):
                if isinstance(value, list):
                    for sub_col_idx, sub_value in enumerate(value, start=col_idx):
                        cell = sheet.cell(row=row_idx, column=sub_col_idx, value=sub_value)
                        if wrap_text:
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    if wrap_text:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(self.file_name)

    def write_data_set_center(self, data, start_column=1, center_columns=None):
        if center_columns is None:
            center_columns = []
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]

        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, value in enumerate(row_data, start=start_column):
                if isinstance(value, list):
                    for sub_col_idx, sub_value in enumerate(value, start=col_idx):
                        cell = sheet.cell(row=row_idx, column=sub_col_idx, value=sub_value)
                        if sub_col_idx in center_columns:
                            cell.alignment = Alignment(horizontal='center')
                else:
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx in center_columns:
                        cell.alignment = Alignment(horizontal='center')

        wb.save(self.file_name)

    def write_module_develop_data(self, data, data2):
        center_columns = [1, 2, 3]
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        no_fill = PatternFill(fill_type=None)

        for row_idx, (row_data, row_data2) in enumerate(zip(data, data2), start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                if col_idx in center_columns:
                    cell.alignment = Alignment(horizontal='center')
                if row_data2[0] == "True":  # Assuming data2 contains a list of lists with one element
                    for col in range(1, len(row_data) + 1):
                        sheet.cell(row=row_idx, column=col).fill = yellow_fill
                else:
                    for col in range(1, len(row_data) + 1):
                        sheet.cell(row=row_idx, column=col).fill = no_fill

        wb.save(self.file_name)

    def clear_merge_column(self, column):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        for merge in list(sheet.merged_cells.ranges):
            if merge.min_col == column:
                sheet.unmerge_cells(str(merge))
        wb.save(self.file_name)

    def merge_cells(self, merge_ranges):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        for merge_range in merge_ranges:
            sheet.merge_cells(merge_range)

        wb.save(self.file_name)

    def write_paradox_data(self, data):
        date_data = [
            ['*为好评率不高于80%,**为好评率不高于50%，***为好评率不高于30%'], [], [], [],
            [' "-" 为该干员没有悖论模拟'], [], [],
            [f'更新日期：{date}']]
        merge_ranges = ['A1:A4', 'A5:A7', 'A8:A10']  # 合并单元格范围
        self.clear_merge_column(1)
        self.write_data_all_center(date_data, wrap_text=True)
        self.merge_cells(merge_ranges)
        for i, sub_list in enumerate(data, start=1):
            self.write_data_all_center(sub_list, start_column=i * 4 - 2)

    def write_module_data(self, data):
        center_columns = [1, 2, 3]
        self.write_data_set_center(data, center_columns=center_columns)

    def write_module_task_data(self, data):
        center_columns = [1, 2]
        self.write_data_set_center(data, center_columns=center_columns)


if __name__ == '__main__':
    # test_data1 = Do_excel("Excel/悖论模拟干员名单作者版.xlsx", "Sheet1").get_data()
    # print(test_data1)
    # test_data2 = Do_excel("Excel/悖论模拟干员名单用户版.xlsx", "Sheet1").get_data()
    # print(test_data2)

    paradox_develop = DoTxt('Auto/Full-Auto/CI/paradox_develop.txt').read_paradox_lines()
    # print(paradox_develop)
    DoExcel("Excel/悖论&模组作业作者版.xlsx", "悖论模拟").write_paradox_data(paradox_develop)

    paradox_user = DoTxt('Auto/Full-Auto/CI/paradox_user.txt').read_paradox_lines()
    # print(paradox_user)
    DoExcel("Excel/悖论&模组作业用户版.xlsx", "悖论模拟").write_paradox_data(paradox_user)

    module_task = DoTxt('Auto/Full-Auto/CI/module.txt').read_module_task_lines()
    # print(module_task)
    DoExcel("Excel/模组任务.xlsx", "Sheet1").write_module_task_data(module_task)

    module_develop, data2 = DoTxt('Auto/Full-Auto/CI/module_develop.txt').read_module_develop_lines()
    # print(module_develop)
    DoExcel("Excel/悖论&模组作业作者版.xlsx", "模组任务").write_module_develop_data(module_develop, data2)

    module_user = DoTxt('Auto/Full-Auto/CI/module_user.txt').read_module_user_lines()
    # print(module_user)
    DoExcel("Excel/悖论&模组作业用户版.xlsx", "模组任务").write_module_data(module_user)
